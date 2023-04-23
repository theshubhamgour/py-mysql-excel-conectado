import mysql.connector
import openpyxl

# Connect to the MySQL database
cnx = mysql.connector.connect(user='root', password='Password@12',
                              host='localhost', database='my_database')

# Open the Excel file containing the SQL queries
wb = openpyxl.load_workbook('queries.xlsx')

# Select the sheet containing the queries
sheet = wb['Sheet1']

# Create a new workbook to store the output
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

# Loop through the rows in the sheet and execute each query
for row in sheet.iter_rows(min_row=2, values_only=True):
    query = row[0]
    if query and 'SELECT' in query.upper():
        cursor = cnx.cursor()
        cursor.execute(query)
        result = cursor.fetchall()

        # Write the query result to the output sheet
        for row_index, row_data in enumerate(result):
            for col_index, col_data in enumerate(row_data):
                output_sheet.cell(row=row_index + 1, column=col_index + 1, value=col_data)

        cnx.commit()
        cursor.close()

# Save the output workbook
output_wb.save('output.xlsx')

# Close the database connection
cnx.close()
